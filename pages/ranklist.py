from flask import Blueprint, request, render_template, send_file, jsonify, session
import io
import json
from datetime import datetime
import traceback

ranklist_bp = Blueprint("ranklist", __name__)

# Metadata for Dashboard
metadata = {
    "title": "Smart Rank List Generator",
    "description": "Create professional rank lists, generate PDF/Excel, and manage student results.",
    "image": ".static/pages/ranklist.jpg"
}

def format_name(name):
    """Format name to proper case"""
    try:
        return str(name).strip().title()
    except:
        return str(name)

def calculate_grade(score, total=100):
    """Calculate grade based on score"""
    try:
        percentage = (float(score) / float(total)) * 100
        if percentage >= 90:
            return "A+"
        elif percentage >= 80:
            return "A"
        elif percentage >= 70:
            return "B+"
        elif percentage >= 60:
            return "B"
        elif percentage >= 50:
            return "C"
        else:
            return "F"
    except:
        return "N/A"

def calculate_percentage(score, total=100):
    try:
        return round((float(score) / float(total)) * 100, 2)
    except:
        return 0

@ranklist_bp.route("/ranklist", methods=["GET", "POST"])
def ranklist_main():
    # Handle GET request - show template
    if request.method == "GET":
        return render_template("ranklist.html")
    
    try:
        # Check if request has JSON data
        if request.is_json:
            data = request.get_json()
            action = data.get('action')
        else:
            action = request.form.get('action')
            data = request.form
        
        # Generate rank list
        if action == 'generate' or (data and 'data' in data):
            # Get form data
            if request.is_json:
                school = data.get('school', '')
                class_name = data.get('class', '')
                entries_data = data.get('data', '')
                exam_name = data.get('exam_name', '')
                max_score = int(data.get('max_score', 100))
            else:
                school = request.form.get('school', '')
                class_name = request.form.get('class', '')
                entries_data = request.form.get('data', '')
                exam_name = request.form.get('exam_name', '')
                max_score = int(request.form.get('max_score', 100))
            
            # Parse entries
            entries = []
            if entries_data:
                raw_entries = entries_data.split(",")
                
                for entry in raw_entries:
                    entry = entry.strip()
                    if not entry:
                        continue
                        
                    if '-' in entry:
                        parts = entry.split('-')
                        name = format_name(parts[0])
                        try:
                            score = float(parts[1]) if parts[1].strip() else 0
                        except:
                            score = 0
                    else:
                        name = format_name(entry)
                        score = 0
                    
                    entries.append({
                        "name": name,
                        "score": score,
                        "percentage": calculate_percentage(score, max_score),
                        "grade": calculate_grade(score, max_score)
                    })
            
            # Sort by score (highest first)
            entries.sort(key=lambda x: x['score'], reverse=True)
            
            # Add rank
            for idx, entry in enumerate(entries):
                entry['rank'] = idx + 1
            
            # Store in session
            session['ranklist_data'] = {
                'school': school,
                'class': class_name,
                'exam_name': exam_name,
                'max_score': max_score,
                'entries': entries,
                'generated_date': datetime.now().strftime("%d-%m-%Y %H:%M:%S")
            }
            
            # Calculate stats
            total_students = len(entries)
            if total_students > 0:
                avg_score = round(sum(e['score'] for e in entries) / total_students, 2)
                highest_score = max(e['score'] for e in entries)
                lowest_score = min(e['score'] for e in entries)
            else:
                avg_score = 0
                highest_score = 0
                lowest_score = 0
            
            return jsonify({
                'success': True,
                'school': school,
                'class': class_name,
                'exam_name': exam_name,
                'max_score': max_score,
                'entries': entries,
                'total_students': total_students,
                'average_score': avg_score,
                'highest_score': highest_score,
                'lowest_score': lowest_score
            })
        
        # Download as Excel (without reportlab)
        elif action == 'download_excel':
            if 'ranklist_data' not in session:
                return jsonify({'error': 'No data found. Please generate rank list first.'}), 400
            
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                data = session['ranklist_data']
                
                # Create Excel workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Rank List"
                
                # Styles
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="1e90ff", end_color="1e90ff", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Add title
                ws.merge_cells('A1:E1')
                ws['A1'] = f"📊 {data['school'] or 'School Name'}"
                ws['A1'].font = Font(bold=True, size=16, color="1e90ff")
                ws['A1'].alignment = Alignment(horizontal="center")
                
                ws.merge_cells('A2:E2')
                ws['A2'] = f"📚 {data['class'] or 'Exam Name'}"
                ws['A2'].font = Font(bold=True, size=12)
                ws['A2'].alignment = Alignment(horizontal="center")
                
                start_row = 4
                
                ws.merge_cells(f'A{start_row}:E{start_row}')
                ws[f'A{start_row}'] = f"🎯 Maximum Score: {data['max_score']}"
                ws[f'A{start_row}'].alignment = Alignment(horizontal="center")
                
                # Headers
                headers = ['Rank', 'Name', 'Score', 'Percentage', 'Grade']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=start_row + 2, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Add data
                for idx, entry in enumerate(data['entries'], start=1):
                    row = start_row + 2 + idx
                    ws.cell(row=row, column=1, value=entry['rank']).border = border
                    ws.cell(row=row, column=2, value=entry['name']).border = border
                    ws.cell(row=row, column=3, value=entry['score']).border = border
                    ws.cell(row=row, column=4, value=f"{entry['percentage']}%").border = border
                    ws.cell(row=row, column=5, value=entry['grade']).border = border
                
                # Adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                return send_file(
                    buffer,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=f"Rank_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                
            except ImportError:
                return jsonify({'error': 'openpyxl not installed. Please install: pip install openpyxl'}), 500
            except Exception as e:
                return jsonify({'error': f'Excel generation error: {str(e)}'}), 500
        
        # Download as PDF (Simple version without reportlab)
        elif action == 'download_pdf':
            if 'ranklist_data' not in session:
                return jsonify({'error': 'No data found. Please generate rank list first.'}), 400
            
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                
                data = session['ranklist_data']
                
                # Create PDF
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
                
                styles = getSampleStyleSheet()
                story = []
                
                # Title
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=24,
                    textColor=colors.HexColor('#1e90ff'),
                    alignment=1,
                    spaceAfter=20
                )
                
                story.append(Paragraph(f"📊 {data['school'] or 'School Name'}", title_style))
                story.append(Paragraph(f"📚 {data['class'] or 'Exam Name'}", styles['Heading2']))
                story.append(Spacer(1, 20))
                
                # Table data
                table_data = [['Rank', 'Name', 'Score', 'Percentage', 'Grade']]
                for entry in data['entries']:
                    table_data.append([
                        str(entry['rank']),
                        entry['name'],
                        str(entry['score']),
                        f"{entry['percentage']}%",
                        entry['grade']
                    ])
                
                # Create table
                table = Table(table_data, repeatRows=1)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e90ff')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                story.append(table)
                story.append(Spacer(1, 20))
                
                # Footer
                footer_style = ParagraphStyle(
                    'FooterStyle',
                    parent=styles['Normal'],
                    fontSize=8,
                    textColor=colors.HexColor('#999999'),
                    alignment=1
                )
                story.append(Paragraph(f"Generated by SK Sinha's Tool Hub | {data['generated_date']}", footer_style))
                
                doc.build(story)
                buffer.seek(0)
                
                return send_file(
                    buffer,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=f"Rank_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                )
                
            except ImportError:
                return jsonify({'error': 'reportlab not installed. Please install: pip install reportlab'}), 500
            except Exception as e:
                return jsonify({'error': f'PDF generation error: {str(e)}'}), 500
        
        else:
            return jsonify({'error': 'Invalid action'}), 400
    
    except Exception as e:
        print(f"Error: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': f'Server error: {str(e)}'}), 500